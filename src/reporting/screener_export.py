import os
import sqlite3

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


class ScreenerExporter:
    """
    Exports screener results to an Excel workbook.
    One worksheet is created for each preset.
    """

    def __init__(self, database_path):
        self.database_path = database_path

    # --------------------------------------------------
    # Load latest financial data
    # --------------------------------------------------

    def load_data(self):

        conn = sqlite3.connect(self.database_path)

        df = pd.read_sql(
            """
            SELECT
                fr.*,
                s.broad_sector
            FROM financial_ratios fr
            LEFT JOIN sectors s
                ON fr.company_id = s.company_id
            """,
            conn
        )

        conn.close()

        # Convert year to numeric
        df["year_num"] = (
            df["year"]
            .astype(str)
            .str[-4:]
            .astype(int)
        )

        # Keep latest year per company
        df = (
            df.sort_values("year_num")
              .drop_duplicates(
                  subset="company_id",
                  keep="last"
              )
              .drop(columns="year_num")
              .reset_index(drop=True)
        )

        return df

        # --------------------------------------------------
    # Export results to Excel
    # --------------------------------------------------

    def export_excel(
        self,
        preset_results,
        output_file
    ):

        os.makedirs(
            os.path.dirname(output_file),
            exist_ok=True
        )

        # ------------------------------------
        # Write worksheets
        # ------------------------------------
        with pd.ExcelWriter(
            output_file,
            engine="openpyxl"
        ) as writer:

            for preset_name, dataframe in preset_results.items():

                dataframe = dataframe.sort_values(
                    by="overall_score",
                    ascending=False
                )

                dataframe.to_excel(
                    writer,
                    sheet_name=preset_name[:31],
                    index=False
                )

        # ------------------------------------
        # Open workbook for formatting
        # ------------------------------------
        workbook = load_workbook(output_file)

        green_fill = PatternFill(
            fill_type="solid",
            start_color="C6EFCE",
            end_color="C6EFCE"
        )

        red_fill = PatternFill(
            fill_type="solid",
            start_color="FFC7CE",
            end_color="FFC7CE"
        )

        # ------------------------------------
        # Format every worksheet
        # ------------------------------------
        for sheet in workbook.worksheets:

            headers = {}

            for cell in sheet[1]:
                headers[cell.value] = cell.column

            if "overall_score" not in headers:
                continue

            score_col = headers["overall_score"]

            # --------------------------------
            # Highlight Overall Score
            # --------------------------------
            for row in range(2, sheet.max_row + 1):

                cell = sheet.cell(
                    row=row,
                    column=score_col
                )

                try:
                    score = float(cell.value)
                except (TypeError, ValueError):
                    continue

                if score >= 70:
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill

            # --------------------------------
            # Auto-fit column widths
            # --------------------------------
            for column_cells in sheet.columns:

                max_length = 0

                for cell in column_cells:

                    if cell.value is not None:

                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                column_letter = get_column_letter(
                    column_cells[0].column
                )

                sheet.column_dimensions[
                    column_letter
                ].width = min(max_length + 3, 35)

        workbook.save(output_file)

        print("\n" + "=" * 60)
        print("SCREENER EXPORT COMPLETED")
        print("=" * 60)
        print(f"Workbook : {output_file}")
        print(f"Sheets   : {len(workbook.sheetnames)}")

        for sheet in workbook.sheetnames:
            print(f"  ✓ {sheet}")

        print("=" * 60)