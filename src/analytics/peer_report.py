import os
import sqlite3

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
DB_PATH = "db/nifty100.db"

OUTPUT_FILE = "output/peer_comparison.xlsx"

GREEN = PatternFill(
    fill_type="solid",
    start_color="C6EFCE",
    end_color="C6EFCE",
)

YELLOW = PatternFill(
    fill_type="solid",
    start_color="FFEB9C",
    end_color="FFEB9C",
)

RED = PatternFill(
    fill_type="solid",
    start_color="FFC7CE",
    end_color="FFC7CE",
)

GOLD = PatternFill(
    fill_type="solid",
    start_color="FFD966",
    end_color="FFD966",
)

BOLD = Font(bold=True)
METRICS = [

    "return_on_equity_pct",
    "return_on_capital_employed",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",

    "debt_to_equity",
    "interest_coverage",
    "asset_turnover",

    "free_cash_flow_cr",
    "cash_from_operations_cr",

    "revenue_cr",
    "net_profit_cr",

    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",

    "earnings_per_share",
    "book_value_per_share",

    "dividend_payout_ratio_pct",

    "capex_intensity_pct",
    "fcf_conversion_pct",

    "composite_quality_score",
]
class PeerComparisonReport:

    def __init__(self):

        self.conn = sqlite3.connect(DB_PATH)

        os.makedirs(
            "output",
            exist_ok=True
        )
    def load_data(self):
        financials = pd.read_sql(
            """
            SELECT *
            FROM financial_ratios
            """,
            self.conn,
        )

        peer_groups = pd.read_sql(
            """
            SELECT
                company_id,
                peer_group_name,
                is_benchmark
            FROM peer_groups
            """,
            self.conn,
        )

        percentiles = pd.read_sql(
            """
            SELECT
                company_id,
                metric,
                percentile_rank
            FROM peer_percentiles
            """,
            self.conn,
        )

        # Keep latest year only
        financials["year_num"] = (
            financials["year"]
            .astype(str)
            .str[-4:]
            .astype(int)
        )

        financials = (
            financials
            .sort_values("year_num")
            .drop_duplicates(
                subset="company_id",
                keep="last",
            )
            .drop(columns="year_num")
            .reset_index(drop=True)
        )

        financials = financials.merge(
            peer_groups,
            on="company_id",
            how="left",
        )

        self.financials = financials
        self.percentiles = percentiles
    def prepare_percentiles(self):

        pivot = self.percentiles.pivot_table(
            index="company_id",
            columns="metric",
            values="percentile_rank"
        )

        pivot.columns = [
            f"{c}_percentile"
            for c in pivot.columns
        ]

        pivot.reset_index(inplace=True)

        self.financials = self.financials.merge(
            pivot,
            on="company_id",
            how="left",
        )
    def generate_excel(self):

        with pd.ExcelWriter(
            OUTPUT_FILE,
            engine="openpyxl"
        ) as writer:

            for peer_group, group in self.financials.groupby("peer_group_name"):

                if pd.isna(peer_group):
                    continue

                sheet = group.copy()

                columns = [
    "company_id",
    "is_benchmark",
]

                columns.extend(METRICS)

                percentile_columns = [
                    f"{m}_percentile"
                    for m in METRICS
                    if f"{m}_percentile" in sheet.columns
                ]

                columns.extend(percentile_columns)

                columns = [
                    c
                    for c in columns
                    if c in sheet.columns
                ]

                sheet = sheet[columns]

                # ---------------------------------
                # Median row
                # ---------------------------------

                median = {}

                median["company_id"] = "Peer Median"
                median["is_benchmark"] = ""

                for m in METRICS:

                    if m in sheet.columns:

                        median[m] = sheet[m].median()

                sheet = pd.concat(
                    [
                        sheet,
                        pd.DataFrame([median])
                    ],
                    ignore_index=True
                )

                sheet.to_excel(
                    writer,
                    sheet_name=peer_group[:31],
                    index=False
                )
    def format_workbook(self):

        workbook = load_workbook(OUTPUT_FILE)

        for sheet in workbook.sheetnames:

            ws = workbook[sheet]

            # -----------------------------
            # Bold Header
            # -----------------------------
            for cell in ws[1]:
                cell.font = BOLD

            headers = [
                cell.value
                for cell in ws[1]
            ]

            # -----------------------------
            # Find percentile columns
            # -----------------------------
            percentile_cols = []

            for i, name in enumerate(headers, start=1):

                if (
                    name is not None
                    and str(name).endswith("_percentile")
                ):
                    percentile_cols.append(i)

            # -----------------------------
            # Colour percentile cells
            # -----------------------------
            for row in range(2, ws.max_row):

                for col in percentile_cols:

                    value = ws.cell(row=row, column=col).value

                    if value is None:
                        continue

                    if value >= 75:
                        ws.cell(row=row, column=col).fill = GREEN

                    elif value <= 25:
                        ws.cell(row=row, column=col).fill = RED

                    else:
                        ws.cell(row=row, column=col).fill = YELLOW

            # -----------------------------
            # Highlight Benchmark Company
            # -----------------------------
            benchmark_col = None

            if "is_benchmark" in headers:
                benchmark_col = headers.index("is_benchmark") + 1

            if benchmark_col:

                for row in range(2, ws.max_row):

                    value = ws.cell(
                        row=row,
                        column=benchmark_col
                    ).value

                    if value == 1:

                        for c in range(1, ws.max_column + 1):

                            ws.cell(
                                row=row,
                                column=c
                            ).fill = GOLD

            # -----------------------------
            # Auto Column Width
            # -----------------------------
            for column_cells in ws.columns:

                length = max(
                    len(str(cell.value))
                    if cell.value is not None
                    else 0
                    for cell in column_cells
                )

                ws.column_dimensions[
                    column_cells[0].column_letter
                ].width = min(length + 3, 30)

        workbook.save(OUTPUT_FILE)

        print(f"\nReport saved to {OUTPUT_FILE}")
if __name__ == "__main__":

    report = PeerComparisonReport()

    report.load_data()

    report.prepare_percentiles()

    report.generate_excel()

    report.format_workbook()

    report.conn.close()
